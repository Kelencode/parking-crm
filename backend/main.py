import asyncio
import io
import json
from contextlib import asynccontextmanager
from datetime import datetime, timedelta
from typing import Annotated, Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from fastapi import Depends, FastAPI, HTTPException, Query, Request, WebSocket, WebSocketDisconnect, status
from fastapi.responses import Response
from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.errors import RateLimitExceeded
from slowapi.util import get_remote_address
from sqlalchemy import text
from fastapi.middleware.cors import CORSMiddleware
from starlette.middleware.base import BaseHTTPMiddleware
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
import bcrypt as _bcrypt
from jose import JWTError, jwt
from pydantic import BaseModel, EmailStr, Field, field_validator
from sqlalchemy import select
from sqlalchemy.orm import selectinload
from sqlalchemy.ext.asyncio import AsyncSession

from database import Base, engine, get_db, settings
from models import (
    AuditAction, AuditLog,
    Incident, IncidentPriority, IncidentStatus, IncidentType,
    JournalEntry, JournalOperation, JournalReason,
    LotType, Software,
    ParkingLot, ShiftNote, ShiftSnapshot, User, UserRole,
)
from utils import calculate_priority, log_action


# ---------- Journal operation / reason maps ----------

_OP_MAP = {
    'въезд': 'entry', 'выезд': 'exit',
    'entry': 'entry', 'exit': 'exit',
}

_REASON_MAP = {
    'Без оплаты(сбой)':         'no_pay_fail',
    'Без оплаты (согласовано)': 'no_pay_approved',
    'Ручное открытие(сбой)':    'manual_fail',
    'Инвалид (перезапись)':     'disabled_rewrite',
    'Инвалид (ручное)':         'disabled_manual',
    'Инвалид(не фРИ\\оплата)': 'disabled_no_fri',
    'Сотрудник ГЦУП(перезапись)': 'staff_rewrite',
    'Сотрудник ГЦУП(ручное)':   'staff_manual',
    'Сотрудник ГЦУП(контролер)': 'staff_controller',
    'Спец. Техника':            'spec_tech',
    'Абонемент':                'abonement',
    'Без карты':                'no_card',
    'Скорая помощь':            'ambulance',
    'Оплата':                   'payment',
    'Электромобиль':            'electric_car',
    'Утеря парковочной карты':  'lost_card',
    'Социальное такси':         'social_taxi',
    'Запись суточного тарифа':  'daily_tariff',
    'Курсус':                   'kursus',
    'Полиция':                  'police',
    'Замятие':                  'jam',
    'Сервис':                   'service',
    'Обслуживание АПС':         'aps_service',
    'Прочее':                   'other',
    'КлеверПарк':               'klever_park',
}
_REASON_REVERSE = {v: k for k, v in _REASON_MAP.items()}
_OP_DISPLAY = {'entry': 'Въезд', 'exit': 'Выезд'}


# ---------- Datetime helpers ----------

def parse_date_from(s: str) -> datetime:
    return datetime.strptime(s, "%Y-%m-%d")

def parse_date_to(s: str) -> datetime:
    return datetime.strptime(s, "%Y-%m-%d").replace(hour=23, minute=59, second=59)

def strip_tz(dt: Optional[datetime]) -> Optional[datetime]:
    if dt is None:
        return None
    if dt.tzinfo is not None:
        from datetime import timezone as _tz
        return dt.astimezone(_tz.utc).replace(tzinfo=None)
    return dt


def _migrate(conn) -> None:
    """Добавляет новые колонки в существующие таблицы (без потери данных)."""
    migrations = [
        "ALTER TABLE users ADD COLUMN is_active BOOLEAN NOT NULL DEFAULT true",
        "ALTER TABLE users ADD COLUMN must_change_password BOOLEAN NOT NULL DEFAULT false",
        "ALTER TABLE users ADD COLUMN created_at TIMESTAMP",
        "ALTER TABLE parking_lots ADD COLUMN is_active BOOLEAN DEFAULT true",
        "ALTER TABLE parking_lots ADD COLUMN notes TEXT",
        # PostgreSQL enum migrations for JournalReason (ignored on SQLite)
        "ALTER TYPE journalreason ADD VALUE IF NOT EXISTS 'Без оплаты(сбой)'",
        "ALTER TYPE journalreason ADD VALUE IF NOT EXISTS 'Без оплаты (согласовано)'",
        "ALTER TYPE journalreason ADD VALUE IF NOT EXISTS 'Ручное открытие(сбой)'",
        "ALTER TYPE journalreason ADD VALUE IF NOT EXISTS 'Инвалид (перезапись)'",
        "ALTER TYPE journalreason ADD VALUE IF NOT EXISTS 'Инвалид (ручное)'",
        "ALTER TYPE journalreason ADD VALUE IF NOT EXISTS 'Инвалид(не фРИ\\оплата)'",
        "ALTER TYPE journalreason ADD VALUE IF NOT EXISTS 'Сотрудник ГЦУП(перезапись)'",
        "ALTER TYPE journalreason ADD VALUE IF NOT EXISTS 'Сотрудник ГЦУП(ручное)'",
        "ALTER TYPE journalreason ADD VALUE IF NOT EXISTS 'Сотрудник ГЦУП(контролер)'",
        "ALTER TYPE journalreason ADD VALUE IF NOT EXISTS 'Спец. Техника'",
        "ALTER TYPE journalreason ADD VALUE IF NOT EXISTS 'Абонемент'",
        "ALTER TYPE journalreason ADD VALUE IF NOT EXISTS 'Без карты'",
        "ALTER TYPE journalreason ADD VALUE IF NOT EXISTS 'Оплата'",
        "ALTER TYPE journalreason ADD VALUE IF NOT EXISTS 'Электромобиль'",
        "ALTER TYPE journalreason ADD VALUE IF NOT EXISTS 'Утеря парковочной карты'",
        "ALTER TYPE journalreason ADD VALUE IF NOT EXISTS 'Социальное такси'",
        "ALTER TYPE journalreason ADD VALUE IF NOT EXISTS 'Запись суточного тарифа'",
        "ALTER TYPE journalreason ADD VALUE IF NOT EXISTS 'Курсус'",
        "ALTER TYPE journalreason ADD VALUE IF NOT EXISTS 'Замятие'",
        "ALTER TYPE journalreason ADD VALUE IF NOT EXISTS 'Сервис'",
        "ALTER TYPE journalreason ADD VALUE IF NOT EXISTS 'Обслуживание АПС'",
        "ALTER TYPE journalreason ADD VALUE IF NOT EXISTS 'Прочее'",
        "ALTER TYPE journalreason ADD VALUE IF NOT EXISTS 'КлеверПарк'",
    ]
    for sql in migrations:
        try:
            conn.execute(text(sql))
        except Exception:
            pass  # колонка уже существует или тип не поддерживается
    try:
        conn.execute(text("UPDATE users SET created_at = CURRENT_TIMESTAMP WHERE created_at IS NULL"))
    except Exception:
        pass


@asynccontextmanager
async def lifespan(app: FastAPI):
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)
        await conn.run_sync(_migrate)
    task = asyncio.create_task(auto_snapshot_task())
    yield
    task.cancel()
    try:
        await task
    except asyncio.CancelledError:
        pass


app = FastAPI(title="Parking CRM", version="1.0.0", lifespan=lifespan)

limiter = Limiter(key_func=get_remote_address)
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # TODO: заменить на реальный домен при деплое на VPS
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


class _SecurityHeadersMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request, call_next):
        response = await call_next(request)
        response.headers["X-Content-Type-Options"] = "nosniff"
        response.headers["X-Frame-Options"] = "DENY"
        response.headers["X-XSS-Protection"] = "1; mode=block"
        return response

app.add_middleware(_SecurityHeadersMiddleware)

oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/auth/token")


# ---------- WebSocket connection manager ----------

class ConnectionManager:
    def __init__(self):
        self._connections: dict[WebSocket, str] = {}

    async def connect(self, ws: WebSocket, role: str) -> None:
        await ws.accept()
        self._connections[ws] = role

    def disconnect(self, ws: WebSocket) -> None:
        self._connections.pop(ws, None)

    async def broadcast_to_techs(self, message: dict) -> None:
        data = json.dumps(message, ensure_ascii=False)
        dead = []
        for ws, role in list(self._connections.items()):
            if role in ("tech", "admin"):
                try:
                    await ws.send_text(data)
                except Exception:
                    dead.append(ws)
        for ws in dead:
            self.disconnect(ws)

    async def broadcast_to_all(self, message: dict) -> None:
        data = json.dumps(message, ensure_ascii=False)
        dead = []
        for ws in list(self._connections.keys()):
            try:
                await ws.send_text(data)
            except Exception:
                dead.append(ws)
        for ws in dead:
            self.disconnect(ws)


manager = ConnectionManager()


# ---------- Auth helpers ----------

def verify_password(plain: str, hashed: str) -> bool:
    return _bcrypt.checkpw(plain.encode(), hashed.encode())


def hash_password(plain: str) -> str:
    return _bcrypt.hashpw(plain.encode(), _bcrypt.gensalt()).decode()


def create_access_token(data: dict) -> str:
    payload = data.copy()
    payload["exp"] = datetime.utcnow() + timedelta(minutes=settings.access_token_expire_minutes)
    return jwt.encode(payload, settings.secret_key, algorithm=settings.algorithm)


async def get_current_user(
    token: Annotated[str, Depends(oauth2_scheme)],
    db: AsyncSession = Depends(get_db),
) -> User:
    credentials_error = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Could not validate credentials",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=[settings.algorithm])
        user_id: int = payload.get("sub")
        if user_id is None:
            raise credentials_error
    except JWTError:
        raise credentials_error
    result = await db.execute(select(User).where(User.id == int(user_id)))
    user = result.scalar_one_or_none()
    if user is None:
        raise credentials_error
    return user


# ---------- Schemas ----------

class Token(BaseModel):
    access_token: str
    token_type: str


class UserCreate(BaseModel):
    name: str
    email: EmailStr
    password: str = Field(min_length=8)
    role: UserRole = UserRole.tech


class UserUpdate(BaseModel):
    name: Optional[str] = None
    role: Optional[UserRole] = None
    is_active: Optional[bool] = None


class PasswordResetIn(BaseModel):
    new_password: str = Field(min_length=8)


class PasswordChangeIn(BaseModel):
    new_password: str = Field(min_length=8)


class UserOut(BaseModel):
    id: int
    name: str
    email: str
    role: UserRole
    is_active: bool = True
    must_change_password: bool = False
    created_at: Optional[datetime] = None
    model_config = {"from_attributes": True}


class ParkingLotCreate(BaseModel):
    name: str
    address: Optional[str] = None
    lot_type: Optional[LotType] = None
    software: Optional[Software] = None
    has_controller: bool = False
    cash_registers_count: int = 1
    has_entry_group: bool = False
    has_exit_group: bool = False


class ParkingLotUpdate(BaseModel):
    is_active: Optional[bool] = None
    notes: Optional[str] = Field(default=None, max_length=500)


class ParkingLotOut(ParkingLotCreate):
    id: int
    is_active: bool = True
    notes: Optional[str] = None
    model_config = {"from_attributes": True}


class IncidentCreate(BaseModel):
    parking_lot_id: int
    type: IncidentType
    description: Optional[str] = Field(default="", max_length=1000)
    priority: IncidentPriority


class IncidentUpdate(BaseModel):
    priority: Optional[IncidentPriority] = None


class CloseIncidentBody(BaseModel):
    resolution: Optional[str] = Field(default=None, max_length=2000)


class IncidentOut(BaseModel):
    id: int
    parking_lot_id: int
    parking_lot: Optional[ParkingLotOut] = None
    type: str
    description: str
    priority: IncidentPriority
    status: IncidentStatus
    resolution: Optional[str] = None
    is_repeat: bool = False
    created_by: int
    assigned_to: Optional[int] = None
    assignee: Optional[UserOut] = None
    created_at: datetime
    closed_at: Optional[datetime] = None
    model_config = {"from_attributes": True}


class ShiftNoteCreate(BaseModel):
    text: str
    expires_at: Optional[datetime] = None


class ShiftNoteOut(BaseModel):
    id: int
    text: str
    expires_at: Optional[datetime]
    created_by: int
    created_at: datetime
    is_active: bool
    model_config = {"from_attributes": True}


class ShiftSnapshotOut(BaseModel):
    id: int
    snapshot_time: datetime
    shift_type: str
    date: str
    incidents_json: str
    notes_json: str
    created_by: Optional[int] = None
    creator: Optional[UserOut] = None
    created_at: datetime
    is_auto: bool
    model_config = {"from_attributes": True}


class AuditLogOut(BaseModel):
    id: int
    user_id: int
    user_name: str
    action: AuditAction
    entity: Optional[str] = None
    entity_id: Optional[int] = None
    details: Optional[str] = None
    created_at: datetime
    model_config = {"from_attributes": True}


class JournalEntryCreate(BaseModel):
    parking_lot_id: int
    operation: str
    grz: str = Field(max_length=20)
    reason: str
    note: Optional[str] = Field(default=None, max_length=300)
    ticket_number: Optional[str] = Field(default=None, max_length=20)
    created_at: Optional[datetime] = None


class JournalEntryUpdate(BaseModel):
    parking_lot_id: Optional[int] = None
    operation: Optional[str] = None
    grz: Optional[str] = Field(default=None, max_length=20)
    reason: Optional[str] = None
    note: Optional[str] = Field(default=None, max_length=300)
    ticket_number: Optional[str] = Field(default=None, max_length=20)
    created_at: Optional[datetime] = None


class JournalEntryOut(BaseModel):
    id: int
    created_at: datetime
    parking_lot_id: int
    parking_lot: Optional[ParkingLotOut] = None
    operation: str
    grz: str
    reason: str
    note: Optional[str] = None
    ticket_number: Optional[str] = None
    created_by: int
    creator: Optional[UserOut] = None
    model_config = {"from_attributes": True, "use_enum_values": True}

    @field_validator('operation', mode='before')
    @classmethod
    def _normalize_op(cls, v):
        raw = v.value if hasattr(v, 'value') else str(v)
        return _OP_DISPLAY.get(raw, _OP_DISPLAY.get(raw.lower(), raw))

    @field_validator('reason', mode='before')
    @classmethod
    def _normalize_reason(cls, v):
        raw = v.value if hasattr(v, 'value') else str(v)
        # Already Russian value from enum — reverse-map English key just in case
        return _REASON_REVERSE.get(raw, raw)


# ---------- Reusable query ----------

def _incident_query():
    return select(Incident).options(
        selectinload(Incident.parking_lot),
        selectinload(Incident.assignee),
    )


async def _make_snapshot(db: AsyncSession, shift_type: str, created_by: Optional[int], is_auto: bool) -> ShiftSnapshot:
    now = datetime.now()
    today_str = now.strftime("%Y-%m-%d")

    inc_result = await db.execute(
        _incident_query().where(Incident.status.in_([
            IncidentStatus.new, IncidentStatus.assigned, IncidentStatus.in_progress,
        ]))
    )
    incidents_data = [
        {
            "id": inc.id,
            "type": inc.type,
            "priority": inc.priority.value if hasattr(inc.priority, "value") else inc.priority,
            "status": inc.status.value if hasattr(inc.status, "value") else inc.status,
            "parking_lot": inc.parking_lot.name if inc.parking_lot else "—",
            "description": inc.description,
        }
        for inc in inc_result.scalars().all()
    ]

    notes_result = await db.execute(
        select(ShiftNote)
        .where(ShiftNote.is_active == True)
        .where((ShiftNote.expires_at == None) | (ShiftNote.expires_at > now))
        .order_by(ShiftNote.created_at.desc())
    )
    notes_data = [
        {"id": n.id, "text": n.text, "created_at": n.created_at.isoformat()}
        for n in notes_result.scalars().all()
    ]

    existing = await db.execute(
        select(ShiftSnapshot).where(
            ShiftSnapshot.date == today_str,
            ShiftSnapshot.shift_type == shift_type,
        )
    )
    snap = existing.scalar_one_or_none()
    if snap:
        snap.snapshot_time  = now
        snap.incidents_json = json.dumps(incidents_data, ensure_ascii=False)
        snap.notes_json     = json.dumps(notes_data, ensure_ascii=False)
        snap.created_by     = created_by
        snap.is_auto        = is_auto
    else:
        snap = ShiftSnapshot(
            snapshot_time=now,
            shift_type=shift_type,
            date=today_str,
            incidents_json=json.dumps(incidents_data, ensure_ascii=False),
            notes_json=json.dumps(notes_data, ensure_ascii=False),
            created_by=created_by,
            is_auto=is_auto,
        )
        db.add(snap)

    await db.commit()
    await db.refresh(snap)
    return snap


async def auto_snapshot_task():
    while True:
        try:
            await asyncio.sleep(60)
            now = datetime.now()
            if (now.hour == 8 and now.minute == 0) or (now.hour == 20 and now.minute == 0):
                shift_type = "day" if now.hour == 8 else "night"
                async with AsyncSession(engine, expire_on_commit=False) as db:
                    await _make_snapshot(db, shift_type, created_by=None, is_auto=True)
        except asyncio.CancelledError:
            break
        except Exception as exc:
            print(f"[auto_snapshot_task] error: {exc}")


# ---------- Auth routes ----------

@app.post("/auth/token", response_model=Token)
@limiter.limit("10/minute")
async def login(request: Request, form: OAuth2PasswordRequestForm = Depends(), db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(User).where(User.email == form.username))
    user = result.scalar_one_or_none()
    if not user or not verify_password(form.password, user.password_hash):
        raise HTTPException(status_code=401, detail="Incorrect email or password")
    if not user.is_active:
        raise HTTPException(status_code=403, detail="Account is disabled")
    token = create_access_token({"sub": str(user.id)})
    await log_action(db, user, AuditAction.login)
    await db.commit()
    return {"access_token": token, "token_type": "bearer"}


@app.get("/auth/me", response_model=UserOut)
async def me(current_user: User = Depends(get_current_user)):
    return current_user


@app.patch("/auth/change-password", status_code=204)
async def change_password(
    data: PasswordChangeIn,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    result = await db.execute(select(User).where(User.id == current_user.id))
    user = result.scalar_one()
    user.password_hash = hash_password(data.new_password)
    user.must_change_password = False
    await log_action(db, current_user, AuditAction.user_updated, "user", current_user.id,
                     {"action": "password_changed"})
    await db.commit()


@app.post("/auth/register", response_model=UserOut, status_code=201)
async def register(data: UserCreate, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    if current_user.role != UserRole.admin:
        raise HTTPException(403, "Only admins can register new users")
    existing = await db.execute(select(User).where(User.email == data.email))
    if existing.scalar_one_or_none():
        raise HTTPException(400, "Email already registered")
    user = User(name=data.name, email=data.email, password_hash=hash_password(data.password), role=data.role)
    db.add(user)
    await db.flush()
    await log_action(db, current_user, AuditAction.user_created, "user", user.id, {"email": data.email, "role": data.role})
    await db.commit()
    await db.refresh(user)
    return user


# ---------- Users ----------

@app.get("/users", response_model=list[UserOut])
async def list_users(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role != UserRole.admin:
        raise HTTPException(403, "Only admins can list users")
    result = await db.execute(select(User).order_by(User.created_at))
    return result.scalars().all()


@app.post("/users", response_model=UserOut, status_code=201)
async def create_user(
    data: UserCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role != UserRole.admin:
        raise HTTPException(403, "Only admins can create users")
    existing = await db.execute(select(User).where(User.email == data.email))
    if existing.scalar_one_or_none():
        raise HTTPException(400, "Email already registered")
    user = User(
        name=data.name,
        email=data.email,
        password_hash=hash_password(data.password),
        role=data.role,
        must_change_password=True,
    )
    db.add(user)
    await db.flush()
    await log_action(db, current_user, AuditAction.user_created, "user", user.id,
                     {"email": data.email, "role": data.role})
    await db.commit()
    await db.refresh(user)
    return user


@app.patch("/users/{user_id}", response_model=UserOut)
async def update_user(
    user_id: int,
    data: UserUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role != UserRole.admin:
        raise HTTPException(403, "Only admins can edit users")
    result = await db.execute(select(User).where(User.id == user_id))
    target = result.scalar_one_or_none()
    if not target:
        raise HTTPException(404, "User not found")

    # Нельзя деактивировать последнего активного admin
    if data.is_active is False or data.role not in (None, UserRole.admin):
        if target.role == UserRole.admin:
            admins = await db.execute(
                select(User).where(User.role == UserRole.admin, User.is_active == True, User.id != user_id)
            )
            if not admins.scalars().first():
                raise HTTPException(400, "Cannot remove the last active admin")

    changes: dict = {}
    if data.name is not None:
        target.name = data.name
        changes["name"] = data.name
    if data.role is not None:
        target.role = data.role
        changes["role"] = data.role
    if data.is_active is not None:
        target.is_active = data.is_active
        changes["is_active"] = data.is_active

    if changes:
        await log_action(db, current_user, AuditAction.user_updated, "user", user_id, changes)
    await db.commit()
    await db.refresh(target)
    return target


@app.post("/users/{user_id}/reset-password", status_code=204)
async def reset_user_password(
    user_id: int,
    data: PasswordResetIn,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role != UserRole.admin:
        raise HTTPException(403, "Only admins can reset passwords")
    result = await db.execute(select(User).where(User.id == user_id))
    target = result.scalar_one_or_none()
    if not target:
        raise HTTPException(404, "User not found")
    target.password_hash = hash_password(data.new_password)
    target.must_change_password = True
    await log_action(db, current_user, AuditAction.user_updated, "user", user_id,
                     {"action": "password_reset"})
    await db.commit()


# ---------- Parking lots ----------

@app.get("/parking-lots", response_model=list[ParkingLotOut])
async def list_parking_lots(db: AsyncSession = Depends(get_db), _: User = Depends(get_current_user)):
    result = await db.execute(select(ParkingLot))
    return result.scalars().all()


@app.patch("/parking-lots/{lot_id}", response_model=ParkingLotOut)
async def update_parking_lot(
    lot_id: int,
    data: ParkingLotUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role not in (UserRole.dispatcher, UserRole.admin):
        raise HTTPException(403, "Only dispatchers/admins can edit parking lots")
    result = await db.execute(select(ParkingLot).where(ParkingLot.id == lot_id))
    lot = result.scalar_one_or_none()
    if not lot:
        raise HTTPException(404, "Parking lot not found")
    for field, value in data.model_dump(exclude_none=True).items():
        setattr(lot, field, value)
    await db.commit()
    await db.refresh(lot)
    return lot


@app.post("/parking-lots", response_model=ParkingLotOut, status_code=201)
async def create_parking_lot(
    data: ParkingLotCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role not in (UserRole.dispatcher, UserRole.admin):
        raise HTTPException(403, "Only dispatchers/admins can manage parking lots")
    lot = ParkingLot(**data.model_dump())
    db.add(lot)
    await db.commit()
    await db.refresh(lot)
    return lot


# ---------- Incidents ----------

@app.get("/incidents/suggest-priority")
async def suggest_priority(
    lot_id: int,
    incident_type: str,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_current_user),
):
    lot_r = await db.execute(select(ParkingLot).where(ParkingLot.id == lot_id))
    lot = lot_r.scalar_one_or_none()
    if not lot:
        raise HTTPException(404, "Parking lot not found")
    return {"priority": calculate_priority(incident_type, lot)}


@app.get("/incidents", response_model=list[IncidentOut])
async def list_incidents(
    status: list[IncidentStatus] = Query(default=[]),
    lot_id: Optional[int] = None,
    date: Optional[str] = None,
    created_date: Optional[str] = None,
    closed_date: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    limit: Optional[int] = None,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_current_user),
):
    query = _incident_query()
    if status:
        query = query.where(Incident.status.in_(status))
    if lot_id:
        query = query.where(Incident.parking_lot_id == lot_id)
    if date:
        # incidents that were active during this date
        query = query.where(Incident.created_at <= parse_date_to(date))
        query = query.where(
            (Incident.closed_at == None) | (Incident.closed_at >= parse_date_from(date))
        )
    if created_date:
        query = query.where(Incident.created_at >= parse_date_from(created_date))
        query = query.where(Incident.created_at <= parse_date_to(created_date))
    if closed_date:
        query = query.where(Incident.closed_at >= parse_date_from(closed_date))
        query = query.where(Incident.closed_at <= parse_date_to(closed_date))
    if date_from:
        query = query.where(Incident.created_at >= parse_date_from(date_from))
    if date_to:
        query = query.where(Incident.created_at <= parse_date_to(date_to))
    query = query.order_by(Incident.created_at.desc())
    if limit:
        query = query.limit(limit)
    result = await db.execute(query)
    return result.scalars().all()


@app.post("/incidents", response_model=IncidentOut, status_code=201)
async def create_incident(
    data: IncidentCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    lot_r = await db.execute(select(ParkingLot).where(ParkingLot.id == data.parking_lot_id))
    lot = lot_r.scalar_one_or_none()
    if not lot:
        raise HTTPException(404, "Parking lot not found")

    incident = Incident(
        parking_lot_id=data.parking_lot_id,
        type=data.type,
        description=data.description or "",
        priority=data.priority,
        created_by=current_user.id,
    )
    db.add(incident)
    await db.flush()
    await log_action(db, current_user, AuditAction.incident_created, "incident", incident.id,
                     {"lot": lot.name, "type": data.type, "priority": data.priority})
    await db.commit()

    await manager.broadcast_to_techs({
        "type": "new_incident",
        "data": {
            "id": incident.id,
            "lot_name": lot.name,
            "description": data.description or data.type,
            "priority": data.priority,
        },
    })

    result = await db.execute(_incident_query().where(Incident.id == incident.id))
    return result.scalar_one()


@app.patch("/incidents/{incident_id}/accept", response_model=IncidentOut)
async def accept_incident(
    incident_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role != UserRole.tech:
        raise HTTPException(403, "Only tech staff can accept incidents")
    result = await db.execute(select(Incident).where(Incident.id == incident_id))
    incident = result.scalar_one_or_none()
    if not incident:
        raise HTTPException(404, "Incident not found")
    if incident.status != IncidentStatus.new:
        raise HTTPException(400, "Only new incidents can be accepted")
    incident.assigned_to = current_user.id
    incident.status = IncidentStatus.in_progress
    await log_action(db, current_user, AuditAction.incident_accepted, "incident", incident_id)
    await db.commit()
    result = await db.execute(_incident_query().where(Incident.id == incident_id))
    return result.scalar_one()


@app.patch("/incidents/{incident_id}/close", response_model=IncidentOut)
async def close_incident(
    incident_id: int,
    body: CloseIncidentBody = CloseIncidentBody(),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    result = await db.execute(select(Incident).where(Incident.id == incident_id))
    incident = result.scalar_one_or_none()
    if not incident:
        raise HTTPException(404, "Incident not found")
    if incident.status == IncidentStatus.closed:
        raise HTTPException(400, "Incident is already closed")
    incident.status = IncidentStatus.closed
    incident.closed_at = datetime.utcnow()
    if body.resolution:
        incident.resolution = body.resolution
    details = {"resolution": body.resolution} if body.resolution else None
    await log_action(db, current_user, AuditAction.incident_closed, "incident", incident_id, details)
    await db.commit()
    result = await db.execute(_incident_query().where(Incident.id == incident_id))
    return result.scalar_one()


@app.patch("/incidents/{incident_id}/reopen", response_model=IncidentOut)
async def reopen_incident(
    incident_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role not in (UserRole.dispatcher, UserRole.admin):
        raise HTTPException(403, "Only dispatchers/admins can reopen incidents")
    result = await db.execute(select(Incident).where(Incident.id == incident_id))
    incident = result.scalar_one_or_none()
    if not incident:
        raise HTTPException(404, "Incident not found")
    if incident.status != IncidentStatus.closed:
        raise HTTPException(400, "Only closed incidents can be reopened")

    lot_r = await db.execute(select(ParkingLot).where(ParkingLot.id == incident.parking_lot_id))
    lot = lot_r.scalar_one_or_none()

    incident.status = IncidentStatus.new
    incident.assigned_to = None
    incident.closed_at = None
    incident.resolution = None
    incident.is_repeat = True
    await log_action(db, current_user, AuditAction.incident_reopened, "incident", incident_id)
    await db.commit()

    if lot:
        await manager.broadcast_to_techs({
            "type": "new_incident",
            "data": {
                "id": incident.id,
                "lot_name": lot.name,
                "description": incident.description or incident.type,
                "priority": incident.priority,
            },
        })

    result = await db.execute(_incident_query().where(Incident.id == incident_id))
    return result.scalar_one()


@app.patch("/incidents/{incident_id}", response_model=IncidentOut)
async def update_incident(
    incident_id: int,
    data: IncidentUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    result = await db.execute(select(Incident).where(Incident.id == incident_id))
    incident = result.scalar_one_or_none()
    if not incident:
        raise HTTPException(404, "Incident not found")
    updates = data.model_dump(exclude_none=True)
    for field, value in updates.items():
        setattr(incident, field, value)
    if updates:
        await log_action(db, current_user, AuditAction.incident_updated, "incident", incident_id, updates)
    await db.commit()
    result = await db.execute(_incident_query().where(Incident.id == incident_id))
    return result.scalar_one()


# ── Report helpers ────────────────────────────────────────────────────────────

def _xlsx_header_fill():
    return PatternFill("solid", fgColor="185FA5")

def _xlsx_row_fill(idx: int):
    return PatternFill("solid", fgColor="EBF3FB") if idx % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")

def _xlsx_header_font():
    return Font(bold=True, color="FFFFFF", size=10)

def _xlsx_auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

PRIORITY_RU = {"critical": "Критичный", "high": "Высокий", "medium": "Средний", "low": "Низкий"}
STATUS_RU   = {"new": "Новая", "in_progress": "В работе", "closed": "Закрыта"}


@app.get("/reports/incidents")
async def report_incidents(
    date_from: str = Query(...),
    date_to:   str = Query(...),
    lot_id:    Optional[int] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role not in (UserRole.dispatcher, UserRole.admin):
        raise HTTPException(403, "Только диспетчеры и администраторы могут скачивать отчёты")

    dt_from = datetime.strptime(date_from, "%Y-%m-%d")
    dt_to   = datetime.strptime(date_to,   "%Y-%m-%d").replace(hour=23, minute=59, second=59)

    query = (
        _incident_query()
        .where(Incident.created_at >= dt_from)
        .where(Incident.created_at <= dt_to)
    )
    if lot_id is not None:
        query = query.where(Incident.parking_lot_id == lot_id)
    result = await db.execute(query.order_by(Incident.created_at))
    incidents = result.scalars().all()

    lot_name = ""
    if lot_id is not None:
        lot_result = await db.execute(select(ParkingLot).where(ParkingLot.id == lot_id))
        lot = lot_result.scalar_one_or_none()
        if lot:
            lot_name = lot.name

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Заявки"
    ws.freeze_panes = "A2"

    headers = ["№", "Дата открытия", "Время", "Стоянка", "Тип сбоя", "Описание",
               "Приоритет", "Статус", "Исполнитель", "Что сделано", "Дата закрытия", "Повторная"]
    ws.append(headers)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = _xlsx_header_fill()
        cell.font = _xlsx_header_font()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28

    for row_idx, inc in enumerate(incidents, 2):
        closed_str = inc.closed_at.strftime("%d.%m.%Y") if inc.closed_at else ""
        row = [
            inc.id,
            inc.created_at.strftime("%d.%m.%Y"),
            inc.created_at.strftime("%H:%M"),
            inc.parking_lot.name if inc.parking_lot else "",
            inc.type,
            inc.description or "",
            PRIORITY_RU.get(inc.priority.value if hasattr(inc.priority, "value") else inc.priority, str(inc.priority)),
            STATUS_RU.get(inc.status.value if hasattr(inc.status, "value") else inc.status, str(inc.status)),
            inc.assignee.name if inc.assignee else "",
            inc.resolution or "",
            closed_str,
            "Да" if inc.is_repeat else "Нет",
        ]
        ws.append(row)
        fill = _xlsx_row_fill(row_idx)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = fill
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    _xlsx_auto_width(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_lot = lot_name.replace(" ", "_") if lot_name else ""
    prefix = f"incidents_{safe_lot}_" if safe_lot else "incidents_"
    filename = f"{prefix}{date_from}_{date_to}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/reports/snapshots")
async def report_snapshots(
    date_from: str = Query(...),
    date_to:   str = Query(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role not in (UserRole.dispatcher, UserRole.admin):
        raise HTTPException(403, "Только диспетчеры и администраторы могут скачивать отчёты")

    result = await db.execute(
        select(ShiftSnapshot)
        .where(ShiftSnapshot.date >= date_from)
        .where(ShiftSnapshot.date <= date_to)
        .options(selectinload(ShiftSnapshot.creator))
        .order_by(ShiftSnapshot.date, ShiftSnapshot.shift_type)
    )
    snaps = result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Срезы смен"
    ws.freeze_panes = "A2"

    headers = ["Дата", "Смена", "Время среза", "Способ", "Зафиксировал",
               "Активных заявок", "Стоянки с заявками"]
    ws.append(headers)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = _xlsx_header_fill()
        cell.font = _xlsx_header_font()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28

    SHIFT_LABEL = {"day": "Дневная 08:00", "night": "Ночная 20:00"}

    for row_idx, snap in enumerate(snaps, 2):
        try:
            incidents = json.loads(snap.incidents_json)
        except Exception:
            incidents = []
        lot_names = list({
            (inc.get("parking_lot") or {}).get("name") or inc.get("parking_lot") or ""
            for inc in incidents
            if (inc.get("parking_lot") or {}).get("name") or inc.get("parking_lot")
        })
        lot_names = [n for n in lot_names if n]

        row = [
            snap.date,
            SHIFT_LABEL.get(snap.shift_type, snap.shift_type),
            snap.snapshot_time.strftime("%H:%M"),
            "Авто" if snap.is_auto else "Вручную",
            snap.creator.name if snap.creator else "Система",
            len(incidents),
            ", ".join(sorted(lot_names)),
        ]
        ws.append(row)
        fill = _xlsx_row_fill(row_idx)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = fill
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    _xlsx_auto_width(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"snapshots_{date_from}_{date_to}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/reports/journal")
async def report_journal(
    date_from: str = Query(...),
    date_to:   str = Query(...),
    lot_id:    Optional[int] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role not in (UserRole.dispatcher, UserRole.admin):
        raise HTTPException(403, "Только диспетчеры и администраторы могут скачивать отчёты")

    query = (
        select(JournalEntry)
        .options(selectinload(JournalEntry.parking_lot), selectinload(JournalEntry.creator))
        .where(JournalEntry.created_at >= parse_date_from(date_from))
        .where(JournalEntry.created_at <= parse_date_to(date_to))
    )
    if lot_id is not None:
        query = query.where(JournalEntry.parking_lot_id == lot_id)
    result = await db.execute(query.order_by(JournalEntry.created_at))
    entries = result.scalars().all()

    lot_name = ""
    if lot_id is not None:
        lot_r = await db.execute(select(ParkingLot).where(ParkingLot.id == lot_id))
        lot = lot_r.scalar_one_or_none()
        if lot:
            lot_name = lot.name

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Журнал"
    ws.freeze_panes = "A2"

    headers = ["№", "Дата", "Время", "Стоянка", "Операция", "ГРЗ",
               "Основание", "Примечание", "Билет", "Кто внёс"]
    ws.append(headers)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = _xlsx_header_fill()
        cell.font = _xlsx_header_font()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28

    ENTRY_FILL = {
        "въезд": PatternFill("solid", fgColor="DBEAFE"),
        "выезд": PatternFill("solid", fgColor="DCFCE7"),
    }

    for row_idx, e in enumerate(entries, 2):
        op = e.operation.value if hasattr(e.operation, "value") else e.operation
        row = [
            e.id,
            e.created_at.strftime("%d.%m.%Y"),
            e.created_at.strftime("%H:%M"),
            e.parking_lot.name if e.parking_lot else "",
            op,
            e.grz,
            e.reason.value if hasattr(e.reason, "value") else e.reason,
            e.note or "",
            e.ticket_number or "",
            e.creator.name if e.creator else "",
        ]
        ws.append(row)
        fill = ENTRY_FILL.get(op, _xlsx_row_fill(row_idx))
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = fill
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    _xlsx_auto_width(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_lot = lot_name.replace(" ", "_") if lot_name else ""
    prefix = f"journal_{safe_lot}_" if safe_lot else "journal_"
    filename = f"{prefix}{date_from}_{date_to}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------- Shift notes ----------

@app.get("/shift-notes", response_model=list[ShiftNoteOut])
async def list_shift_notes(
    date: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_current_user),
):
    if date:
        result = await db.execute(
            select(ShiftNote)
            .where(ShiftNote.created_at >= parse_date_from(date))
            .where(ShiftNote.created_at <= parse_date_to(date))
            .order_by(ShiftNote.created_at.desc())
        )
    else:
        now = datetime.utcnow()
        result = await db.execute(
            select(ShiftNote)
            .where(ShiftNote.is_active == True)
            .where((ShiftNote.expires_at == None) | (ShiftNote.expires_at > now))
            .order_by(ShiftNote.created_at.desc())
        )
    return result.scalars().all()


@app.post("/shift-notes", response_model=ShiftNoteOut, status_code=201)
async def create_shift_note(
    data: ShiftNoteCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    dump = data.model_dump()
    dump["expires_at"] = strip_tz(dump.get("expires_at"))
    note = ShiftNote(**dump, created_by=current_user.id)
    db.add(note)
    await db.flush()
    await log_action(db, current_user, AuditAction.shift_note_created, "shift_note", note.id)
    await db.commit()
    await db.refresh(note)
    return note


@app.delete("/shift-notes/{note_id}", status_code=204)
async def deactivate_shift_note(
    note_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    result = await db.execute(select(ShiftNote).where(ShiftNote.id == note_id))
    note = result.scalar_one_or_none()
    if not note:
        raise HTTPException(404, "Note not found")
    note.is_active = False
    await log_action(db, current_user, AuditAction.shift_note_deleted, "shift_note", note_id)
    await db.commit()


# ---------- Shift snapshots ----------

@app.post("/shift-snapshots", response_model=ShiftSnapshotOut, status_code=201)
async def create_shift_snapshot(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role not in (UserRole.dispatcher, UserRole.admin):
        raise HTTPException(403, "Only dispatchers/admins can create snapshots")
    now = datetime.now()
    shift_type = "day" if 6 <= now.hour < 18 else "night"
    snap = await _make_snapshot(db, shift_type, created_by=current_user.id, is_auto=False)
    result = await db.execute(
        select(ShiftSnapshot)
        .options(selectinload(ShiftSnapshot.creator))
        .where(ShiftSnapshot.id == snap.id)
    )
    return result.scalar_one()


@app.get("/shift-snapshots/latest", response_model=ShiftSnapshotOut)
async def get_latest_snapshot(
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_current_user),
):
    result = await db.execute(
        select(ShiftSnapshot)
        .options(selectinload(ShiftSnapshot.creator))
        .order_by(ShiftSnapshot.created_at.desc())
        .limit(1)
    )
    snap = result.scalar_one_or_none()
    if not snap:
        raise HTTPException(404, "No snapshots yet")
    return snap


@app.get("/shift-snapshots", response_model=list[ShiftSnapshotOut])
async def get_shift_snapshots(
    date: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_current_user),
):
    if not date:
        date = datetime.now().strftime("%Y-%m-%d")
    result = await db.execute(
        select(ShiftSnapshot)
        .options(selectinload(ShiftSnapshot.creator))
        .where(ShiftSnapshot.date == date)
        .order_by(ShiftSnapshot.shift_type)
    )
    return result.scalars().all()


# ---------- Journal ----------

@app.get("/journal", response_model=list[JournalEntryOut])
async def list_journal(
    date_from:      Optional[str]              = None,
    date_to:        Optional[str]              = None,
    parking_lot_id: Optional[int]              = None,
    operation:      Optional[JournalOperation] = None,
    reason:         Optional[JournalReason]    = None,
    limit:          int                        = Query(default=200, le=1000),
    offset:         int                        = 0,
    db: AsyncSession = Depends(get_db),
    _:  User         = Depends(get_current_user),
):
    query = (
        select(JournalEntry)
        .options(selectinload(JournalEntry.parking_lot), selectinload(JournalEntry.creator))
    )
    if date_from:
        query = query.where(JournalEntry.created_at >= datetime.strptime(date_from, "%Y-%m-%d"))
    if date_to:
        dt_to = datetime.strptime(date_to, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
        query = query.where(JournalEntry.created_at <= dt_to)
    if parking_lot_id:
        query = query.where(JournalEntry.parking_lot_id == parking_lot_id)
    if operation:
        query = query.where(JournalEntry.operation == operation)
    if reason:
        query = query.where(JournalEntry.reason == reason)
    query = query.order_by(JournalEntry.created_at.desc()).limit(limit).offset(offset)
    result = await db.execute(query)
    return result.scalars().all()


@app.post("/journal", response_model=JournalEntryOut, status_code=201)
async def create_journal_entry(
    data: JournalEntryCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role not in (UserRole.dispatcher, UserRole.admin):
        raise HTTPException(403, "Только диспетчеры и администраторы могут создавать записи")

    lot_r = await db.execute(select(ParkingLot).where(ParkingLot.id == data.parking_lot_id))
    lot = lot_r.scalar_one_or_none()
    if not lot:
        raise HTTPException(404, "Parking lot not found")

    created_at_clean = strip_tz(data.created_at) if data.created_at else datetime.utcnow()
    operation = _OP_MAP.get(data.operation.lower(), 'exit')
    reason = _REASON_MAP.get(data.reason, data.reason)

    ins = await db.execute(
        text("""
            INSERT INTO journal_entries
                (created_at, parking_lot_id, operation, grz, reason,
                 note, ticket_number, created_by)
            VALUES
                (:created_at, :parking_lot_id, :operation, :grz, :reason,
                 :note, :ticket_number, :created_by)
            RETURNING id
        """),
        {
            "created_at": created_at_clean,
            "parking_lot_id": data.parking_lot_id,
            "operation": operation,
            "grz": data.grz.strip().upper(),
            "reason": reason,
            "note": data.note or None,
            "ticket_number": data.ticket_number or None,
            "created_by": current_user.id,
        },
    )
    await db.commit()
    entry_id = ins.scalar()

    result = await db.execute(
        select(JournalEntry)
        .options(selectinload(JournalEntry.parking_lot), selectinload(JournalEntry.creator))
        .where(JournalEntry.id == entry_id)
    )
    entry_out = result.scalar_one()

    await manager.broadcast_to_all({
        "type": "new_journal_entry",
        "data": {
            "id": entry_out.id,
            "lot_name": lot.name,
            "operation": data.operation,
            "grz": entry_out.grz,
            "created_by": current_user.id,
        },
    })

    return entry_out


@app.delete("/journal/{entry_id}", status_code=204)
async def delete_journal_entry(
    entry_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role != UserRole.admin:
        raise HTTPException(403, "Только администраторы могут удалять записи")
    result = await db.execute(select(JournalEntry).where(JournalEntry.id == entry_id))
    entry = result.scalar_one_or_none()
    if not entry:
        raise HTTPException(404, "Journal entry not found")
    await db.delete(entry)
    await db.commit()


@app.patch("/journal/{entry_id}", response_model=JournalEntryOut)
async def update_journal_entry(
    entry_id: int,
    data: JournalEntryUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role not in (UserRole.dispatcher, UserRole.admin):
        raise HTTPException(403, "Только диспетчеры и администраторы могут редактировать записи")
    result = await db.execute(
        select(JournalEntry)
        .options(selectinload(JournalEntry.parking_lot), selectinload(JournalEntry.creator))
        .where(JournalEntry.id == entry_id)
    )
    entry = result.scalar_one_or_none()
    if not entry:
        raise HTTPException(404, "Journal entry not found")
    updates = data.model_dump(exclude_none=True)
    if "created_at" in updates:
        updates["created_at"] = strip_tz(updates["created_at"])
    if "grz" in updates:
        updates["grz"] = updates["grz"].strip().upper()
    if "operation" in updates:
        updates["operation"] = _OP_MAP.get(updates["operation"].lower(), 'exit')
    if "reason" in updates:
        updates["reason"] = _REASON_MAP.get(updates["reason"], updates["reason"])
    if updates:
        set_clause = ", ".join(f"{k} = :{k}" for k in updates)
        updates["_id"] = entry_id
        await db.execute(
            text(f"UPDATE journal_entries SET {set_clause} WHERE id = :_id"),
            updates,
        )
        await db.commit()
    result = await db.execute(
        select(JournalEntry)
        .options(selectinload(JournalEntry.parking_lot), selectinload(JournalEntry.creator))
        .where(JournalEntry.id == entry_id)
    )
    return result.scalar_one()


# ---------- WebSocket ----------

@app.websocket("/ws")
async def websocket_endpoint(ws: WebSocket, token: str = Query(default=None)):
    if not token:
        await ws.close(code=4001)
        return
    try:
        payload = jwt.decode(token, settings.secret_key, algorithms=[settings.algorithm])
        user_id = payload.get("sub")
        if not user_id:
            await ws.close(code=4001)
            return
    except JWTError:
        await ws.close(code=4001)
        return

    async with AsyncSession(engine, expire_on_commit=False) as db:
        result = await db.execute(select(User).where(User.id == int(user_id)))
        user = result.scalar_one_or_none()
        if not user or not user.is_active:
            await ws.close(code=4001)
            return

    await manager.connect(ws, user.role.value)
    try:
        while True:
            await ws.receive_text()
    except WebSocketDisconnect:
        manager.disconnect(ws)


# ---------- Audit logs ----------

@app.get("/audit-logs", response_model=list[AuditLogOut])
async def list_audit_logs(
    user_id: Optional[int] = None,
    action: Optional[AuditAction] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    limit: int = Query(default=200, le=1000),
    offset: int = 0,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role != UserRole.admin:
        raise HTTPException(403, "Only admins can view audit logs")
    query = select(AuditLog)
    if user_id:
        query = query.where(AuditLog.user_id == user_id)
    if action:
        query = query.where(AuditLog.action == action)
    if date_from:
        query = query.where(AuditLog.created_at >= parse_date_from(date_from))
    if date_to:
        query = query.where(AuditLog.created_at <= parse_date_to(date_to))
    query = query.order_by(AuditLog.created_at.desc()).limit(limit).offset(offset)
    result = await db.execute(query)
    return result.scalars().all()
